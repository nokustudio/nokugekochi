import os
import re
import json
import html
import openpyxl
from PIL import Image
from projects import PROJECTS

TEMPLATE_FILE = "walkthrough-template.html"

# Room renders / product photos are photographic, so lossless PNG buys nothing
# but 5-10x the file size over a high-quality JPEG. Caps are sized well above
# what the lightbox modal ever displays (max-width:1200px CSS, so ~2400px covers
# a 2x-retina screen) — most source images are already under these, so most
# files end up simply recompressed at their native resolution, not downscaled.
ROOM_RENDER_MAX_W = 2600
PRODUCT_IMAGE_MAX_W = 2200
WEB_IMAGE_QUALITY = 92


def optimize_web_image(path, max_w):
    """Convert a lossless PNG render/product photo to a compressed JPEG in place,
    if it isn't one already. The original is preserved as '<path>.orig' (never
    deleted — gitignored, invisible to this script's own directory scans since
    they only match .png/.jpg/.jpeg). Returns the filename to use going forward
    (unchanged if no conversion happened).

    Skips PNGs with real (non-opaque) transparency rather than risk mangling a
    genuine cutout image — those need a human decision, not silent flattening.
    """
    if not path.lower().endswith(".png"):
        return os.path.basename(path)

    im = Image.open(path)
    if im.mode in ("RGBA", "LA"):
        alpha = im.convert("RGBA").split()[-1]
        if alpha.getextrema()[0] < 250:
            print(f"  (skipping optimization — has real transparency) {os.path.basename(path)}")
            return os.path.basename(path)
        im = im.convert("RGB")
    elif im.mode in ("P", "CMYK"):
        im = im.convert("RGB")

    w, h = im.size
    if w > max_w:
        im = im.resize((max_w, round(h * max_w / w)), Image.LANCZOS)

    base, _ = os.path.splitext(path)
    jpg_path = base + ".jpg"
    orig_backup = path + ".orig"
    before = os.path.getsize(path)

    im.save(jpg_path, "JPEG", quality=WEB_IMAGE_QUALITY, optimize=True, progressive=True)
    if os.path.exists(orig_backup):
        os.remove(orig_backup)  # superseded backup (this png replaced an earlier version)
    os.rename(path, orig_backup)

    print(f"  optimized: {os.path.basename(path)} ({before/1024/1024:.2f}MB -> "
          f"{os.path.getsize(jpg_path)/1024/1024:.2f}MB)")
    return os.path.basename(jpg_path)


def optimize_images_in(dir_path, max_w):
    """Run optimize_web_image() on every .png directly inside dir_path (non-recursive —
    callers apply this separately to a room folder and its Products/ subfolder,
    since each needs its own size cap)."""
    if not os.path.isdir(dir_path):
        return
    for item in os.listdir(dir_path):
        item_path = os.path.join(dir_path, item)
        if os.path.isfile(item_path) and item.lower().endswith(".png"):
            optimize_web_image(item_path, max_w)


def list_images_ordered(dir_path):
    """List the image files directly inside dir_path, in slider/display order.

    By default that's just alphabetical. Drop an `order.txt` in the folder to
    override it: one filename per line, in the order you want them to appear.
    Files not mentioned in order.txt are appended afterward, alphabetically —
    so newly-added images still show up without editing the list. Blank lines
    and lines starting with '#' are ignored.
    """
    if not os.path.isdir(dir_path):
        return []
    all_images = sorted(
        item for item in os.listdir(dir_path)
        if os.path.isfile(os.path.join(dir_path, item)) and item.lower().endswith(('.png', '.jpg', '.jpeg'))
    )

    order_path = os.path.join(dir_path, "order.txt")
    if not os.path.exists(order_path):
        return all_images

    with open(order_path, 'r', encoding='utf-8') as f:
        wanted = [line.strip() for line in f if line.strip() and not line.strip().startswith('#')]

    available = set(all_images)
    ordered = [name for name in wanted if name in available]
    remaining = sorted(available - set(ordered))
    return ordered + remaining


def normalize_name(name):
    n = name.lower()
    n = re.sub(r'\.png$|\.jpg$|\.jpeg$', '', n)
    n = re.sub(r'\[.*?\]', '', n) # Strip [SIZE-UP-OPT] internal-code bracket and contents
    n = re.sub(r'\(.*?\)', '', n) # Strip parentheses and contents
    n = re.sub(r'\bv2\b|\bv1\b|\bfull image\b|\bcopy\b', '', n)
    n = re.sub(r'[^a-z0-9]', '', n)
    return n


def merge_rooms_into_shell(shell_html, ordered_blocks, js_room_renders, active_indices):
    """Splice generated room-section blocks + lightbox JS data into the static shell.
    Returns the merged HTML, or None (after printing an error) if a marker is missing."""
    orig_html = shell_html

    # Splice right after the app-container opens. Anything between here and the
    # lightbox (the old header + baked-in rooms) is discarded and replaced — so
    # the removed nav header simply never reaches the output.
    anchor = '<div class="app-container" id="app-main">'
    idx = orig_html.find(anchor)
    if idx == -1:
        print("Error: Could not find app-container anchor.")
        return None
    header_end = idx + len(anchor)

    lightbox_start = orig_html.find('<!-- ==========================================================================\n       FULLSCREEN CINEMATIC LIGHTBOX MODAL')
    if lightbox_start == -1:
        lightbox_start = orig_html.find('<!-- ==========================================================================\r\n       FULLSCREEN CINEMATIC LIGHTBOX MODAL')
    if lightbox_start == -1:
        print("Error: Could not find lightbox start.")
        return None

    gen_rooms_str = "\n".join(ordered_blocks)
    # Room/product imagery loads lazily (hundreds of images per walkthrough);
    # cover images live in the shell and stay eager.
    gen_rooms_str = gen_rooms_str.replace('<img src="', '<img loading="lazy" src="')
    scroll_top_html = """
  <!-- Scroll to Top Button -->
  <button id="scroll-top-btn" onclick="window.scrollTo({top:0,behavior:'smooth'})" title="Back to top">
    <svg viewBox="0 0 24 24"><polyline points="18 15 12 9 6 15"/></svg>
  </button>

  <!-- Home: back to the Good Earth Kochi hub -->
  <a id="home-pill" href="index.html" title="All projects — Good Earth Kochi">
    <svg viewBox="0 0 24 24"><path d="M3 11.5 12 4l9 7.5"/><path d="M5.5 10.5V20h13v-9.5"/></svg>
  </a>

  """
    middle_section = f"""\n\n    <!-- Room Snap Scroll Container -->
    <div class="scroll-container">
{gen_rooms_str}
    </div>
  </div>\n\n  """
    orig_html = orig_html[:header_end] + middle_section + scroll_top_html + orig_html[lightbox_start:]

    script_start = orig_html.find('<script>')
    if script_start == -1:
        print("Error: Could not find script start.")
        return None
    script_start += 8

    lightbox_group_start = orig_html.find('let currentLightboxGroup = null;')
    if lightbox_group_start == -1:
        print("Error: Could not find lightbox group start in javascript.")
        return None

    js_data = f"""const roomRenders = {json.dumps(js_room_renders, indent=2)};\n\nconst activeRenderIndex = {json.dumps(active_indices, indent=2)};\n"""
    js_inject = f"""
    {js_data}
    function scrollToRoom(roomId) {{
      const target = document.getElementById(roomId);
      if (target) {{
        target.scrollIntoView({{ behavior: 'smooth' }});
      }}
    }}

    function highlightRoom(roomId) {{
      const row = document.getElementById('index-row-' + roomId);
      if (row) {{
        row.classList.add('highlighted');
      }}
      const rect = document.getElementById('svg-rect-' + roomId);
      if (rect) {{
        rect.classList.add('highlighted');
      }}
    }}

    function unhighlightRoom(roomId) {{
      const row = document.getElementById('index-row-' + roomId);
      if (row) {{
        row.classList.remove('highlighted');
      }}
      const rect = document.getElementById('svg-rect-' + roomId);
      if (rect) {{
        rect.classList.remove('highlighted');
      }}
    }}

    document.addEventListener('DOMContentLoaded', () => {{
      // Dynamic scroll enabling for rooms with more than 4 products
      document.querySelectorAll('.room-section').forEach(room => {{
        const productGrid = room.querySelector('.products-grid');
        if (productGrid) {{
          const productCount = productGrid.querySelectorAll('.product-card').length;
          const rightCol = room.querySelector('.right-column');
          if (rightCol && productCount > 4) {{
            rightCol.classList.add('has-scroll');
          }}
        }}
      }});

      // Scroll-to-top button visibility: show when user scrolls past the first section
      const scrollTopBtn = document.getElementById('scroll-top-btn');
      if (scrollTopBtn) {{
        window.addEventListener('scroll', () => {{
          // Show if scrolled past the first snap section (~80% of viewport height)
          const threshold = window.innerHeight * 0.8;
          if (window.scrollY > threshold) {{
            scrollTopBtn.classList.add('visible');
          }} else {{
            scrollTopBtn.classList.remove('visible');
          }}
        }});
      }}

      // Floor sliders (Plan Key + Furniture Layout): swipe to change floors
      document.querySelectorAll('.floor-track').forEach((track) => {{
        const prefix = track.id.replace(/-track$/, '');
        const slider = track.closest('.floor-slider');
        if (!slider) return;
        let x0 = null;
        slider.addEventListener('touchstart', (e) => {{ x0 = e.touches[0].clientX; }}, {{ passive: true }});
        slider.addEventListener('touchend', (e) => {{
          if (x0 === null) return;
          const dx = e.changedTouches[0].clientX - x0;
          if (Math.abs(dx) > 40) floorSlide(prefix, dx < 0 ? 1 : -1);
          x0 = null;
        }});
      }});
    }});

    // Floor slider controls — a transform track (deterministic; no scroll-snap fights).
    // Keyed by prefix so the Plan Key ('plan') and Furniture Layout ('furn') sliders
    // on the same page track their own position independently.
    var floorIndex = {{}};
    function floorApply(prefix) {{
      const track = document.getElementById(prefix + '-track');
      if (!track) return;
      const slides = track.querySelectorAll('.floor-slide');
      if (!slides.length) return;
      let i = Math.max(0, Math.min(floorIndex[prefix] || 0, slides.length - 1));
      floorIndex[prefix] = i;
      track.style.transform = 'translateX(' + (-i * 100) + '%)';
      const labelEl = document.getElementById(prefix + '-slider-label');
      if (labelEl) labelEl.textContent = slides[i].dataset.label || '';
      document.querySelectorAll('[data-dots="' + prefix + '"] .floor-dot').forEach((d, j) => d.classList.toggle('active', j === i));
    }}
    function floorSlide(prefix, dir) {{ floorIndex[prefix] = (floorIndex[prefix] || 0) + dir; floorApply(prefix); }}
    function floorGo(prefix, i) {{ floorIndex[prefix] = i; floorApply(prefix); }}

    """
    orig_html = orig_html[:script_start] + js_inject + orig_html[lightbox_group_start:]
    return orig_html


def set_shell_tokens(orig_html, title, meta_description, brand_name):
    """Apply the tokens common to every generated page: <title>, meta description, header brand."""
    orig_html = re.sub(r'<title>.*?</title>',
                       f'<title>{title}</title>', orig_html, count=1, flags=re.S)
    orig_html = re.sub(r'(<meta name="description" content=").*?(")',
                       lambda m: m.group(1) + meta_description + m.group(2),
                       orig_html, count=1, flags=re.S)
    orig_html = re.sub(r'(<span class="brand-name">).*?(</span>)',
                       lambda m: m.group(1) + brand_name + m.group(2), orig_html, count=1, flags=re.S)
    return orig_html


def _cover_bg_html(hero):
    """Build the cover background block from hero config: kenburns | triptych."""
    style = hero.get("cover_style")
    images = hero.get("cover_images", [])
    if style == "kenburns" and images:
        imgs = "\n      ".join(f'<img src="{src}" alt="">' for src in images)
        return f"""<div class="l-bg-kenburns" aria-hidden="true">
      {imgs}
    </div>"""
    if style == "triptych" and images:
        panels = "\n      ".join(
            f'<div class="l-bg-triptych__panel"><img src="{src}" alt=""></div>' for src in images)
        return f"""<div class="l-bg-triptych" aria-hidden="true">
      {panels}
    </div>"""
    return None  # keep the template's default collage


def _cover_facts_html(facts):
    """Staggered fact strip: fact · rule · fact · rule · fact"""
    parts = []
    for i, fact in enumerate(facts):
        delay = 1.65 + i * 0.12   # lands between the descriptor (1.5s) and theme chip (1.9s)
        if i:
            parts.append(f'<span class="l-fact-rule" style="animation-delay:{delay:.2f}s"></span>')
        parts.append(f'<span class="l-fact" style="animation-delay:{delay:.2f}s">{fact}</span>')
    return "".join(parts)


def set_embedded_cover(orig_html, hero, side_ticker_project):
    """Render the per-project cover: hero copy tokens, researched fact strip,
    and the project's background (render slideshow or built-photo triptych)."""
    def _set_tag(html_str, cls, tag, value):
        return re.sub(rf'(<{tag} class="{re.escape(cls)}"[^>]*>).*?(</{tag}>)',
                      lambda m: m.group(1) + value + m.group(2), html_str, count=1, flags=re.S)

    orig_html = _set_tag(orig_html, "l-hero__article", "span", hero["article"])
    orig_html = _set_tag(orig_html, "l-hero__name", "h1", hero["name"])
    orig_html = _set_tag(orig_html, "l-hero__subtitle", "span", hero["subtitle"])
    orig_html = _set_tag(orig_html, "l-hero__eyebrow-text", "span", hero["eyebrow"])
    orig_html = _set_tag(orig_html, "l-theme-chip__text", "span", hero["theme_chip"])
    orig_html = _set_tag(orig_html, "l-room-badge", "span", hero["room_count"])
    orig_html = orig_html.replace(
        "Good Earth Kochi · Umang Residence · Furnishing Proposal",
        f"Good Earth Kochi · {side_ticker_project} · Furnishing Proposal", 1)

    bg_html = _cover_bg_html(hero)
    if bg_html:
        orig_html = re.sub(r'<!-- COVER_BG_START -->.*?<!-- COVER_BG_END -->',
                           bg_html, orig_html, count=1, flags=re.S)

    facts = hero.get("facts")
    if facts:
        orig_html = orig_html.replace("<!-- COVER_FACTS -->", _cover_facts_html(facts), 1)
    return orig_html


MATERIALS_START = "<!-- MATERIALS_MOODBOARD_START -->"
MATERIALS_END = "<!-- MATERIALS_MOODBOARD_END -->"


def inject_materials_section(base_dir):
    """Materials & Moodboard is identical for every project, so it lives once on the
    shared hub (index.html) instead of being repeated inside each project's walkthrough.
    We extract these sections dynamically from walkthrough-template.html."""
    print("\n=== Injecting shared Materials & Moodboard section into index.html ===")

    template_path = os.path.join(base_dir, "walkthrough-template.html")
    if not os.path.exists(template_path):
        print(f"Error: Template file not found at {template_path}")
        return

    with open(template_path, 'r', encoding='utf-8') as f:
        template_content = f.read()

    # Extract materials HTML
    m_start_marker = "<!-- MATERIALS_TEMPLATE_START -->"
    m_end_marker = "<!-- MATERIALS_TEMPLATE_END -->"
    m_start_idx = template_content.find(m_start_marker)
    m_end_idx = template_content.find(m_end_marker)
    if m_start_idx == -1 or m_end_idx == -1:
        print("Error: Could not find MATERIALS_TEMPLATE markers in walkthrough-template.html")
        return
    materials_html = template_content[m_start_idx + len(m_start_marker):m_end_idx].strip()

    # Extract moodboard HTML
    mb_start_marker = "<!-- MOODBOARD_TEMPLATE_START -->"
    mb_end_marker = "<!-- MOODBOARD_TEMPLATE_END -->"
    mb_start_idx = template_content.find(mb_start_marker)
    mb_end_idx = template_content.find(mb_end_marker)
    if mb_start_idx == -1 or mb_end_idx == -1:
        print("Error: Could not find MOODBOARD_TEMPLATE markers in walkthrough-template.html")
        return
    moodboard_html = template_content[mb_start_idx + len(mb_start_marker):mb_end_idx].strip()

    # The hub has no lightbox modal, so drop the click-to-zoom hooks for this embed.
    materials_html = re.sub(r' onclick="openLightbox\([^"]*\)"', '', materials_html)
    moodboard_html = re.sub(r' onclick="openLightbox\([^"]*\)"', '', moodboard_html)

    section_html = f"""<!-- =====================================================================
     SECTION 3: MATERIALS & MOODBOARD — LIGHT THEME (shared across projects)
     ===================================================================== -->
<section class="quote-section" id="materials-moodboard" aria-label="Materials and Moodboard">

  <nav class="qs-nav" aria-label="Section navigation">
    <div class="qs-nav__brand">
      <img src="Noku mark.png" alt="Noku Studio" class="qs-nav__logo">
      <span class="qs-nav__wordmark">Noku Studio</span>
    </div>
    <span class="qs-nav__label">Brief &nbsp;·&nbsp; Two</span>
  </nav>

  <div class="materials-section__stage">
    <div class="materials-section__col">
      <h3 class="materials-section__heading">Materials</h3>
{materials_html}
    </div>
    <div class="materials-section__col">
      <h3 class="materials-section__heading">Moodboard</h3>
{moodboard_html}
    </div>
  </div>

  <div class="qs-footer">
    <span class="qs-pagination">03 &nbsp;/&nbsp; 30</span>
  </div>

</section>"""

    index_path = os.path.join(base_dir, "index.html")
    with open(index_path, 'r', encoding='utf-8') as f:
        index_html = f.read()

    start = index_html.find(MATERIALS_START)
    end = index_html.find(MATERIALS_END)
    if start == -1 or end == -1:
        print("Error: Could not find MATERIALS_MOODBOARD markers in index.html.")
        return

    # Swatch/moodboard images are below the fold on the hub — lazy-load them
    section_html = section_html.replace('<img src="', '<img loading="lazy" src="')

    index_html = (index_html[:start] + MATERIALS_START + "\n" + section_html + "\n"
                 + index_html[end:])

    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(index_html)
    print("Success! Materials & Moodboard section injected into index.html.")


def build_project(project, shell_html, base_dir):
    print(f"\n=== Building project: {project['name']} ===")

    # Per-project config (locals mirror the old hardcoded names so the body stays intact)
    proj_base = project["base"]                       # e.g. "projects/umang"
    proj_dir = os.path.join(base_dir, proj_base)      # absolute disk path
    rooms_config = project["rooms_config"]
    custom_rates_mapping = project["custom_rates_mapping"]
    layout_images_cfg = project["layout_images"]
    copy = project["copy"]

    # Plan-key and furniture layout are per-floor lists (one floor for Umang, two for
    # Modern Times). Keep only floors whose image actually exists on disk, so the site
    # degrades gracefully before content lands.
    keyplan_floors = [f for f in project["keyplan"]["floors"]
                      if os.path.exists(os.path.join(proj_dir, f["image"]))]
    furniture_floors = [f for f in project["furniture_layout"]["floors"]
                        if os.path.exists(os.path.join(proj_dir, f["file"]))]
    has_plan = bool(keyplan_floors)
    has_furniture = bool(furniture_floors)

    # room id -> the floor it belongs to (for the per-room key-plan thumbnail)
    room_floor = {}
    for fl in keyplan_floors:
        for rid in fl["hotspots"]:
            room_floor[rid] = fl

    # 1. Path to the unified Product_Database.xlsx file
    excel_path = os.path.normpath(os.path.join(proj_dir, project["xlsx"]["database"]))

    if not os.path.exists(excel_path):
        print(f"Error: Unified database Excel file not found at: {excel_path}")
        return

    # 2. Get list of room directories (looking for folders starting with digits)
    room_dirs = []
    if os.path.isdir(proj_dir):
        for item in sorted(os.listdir(proj_dir)):
            item_path = os.path.join(proj_dir, item)
            if os.path.isdir(item_path) and re.match(r'^\d+\.', item):
                room_dirs.append(item)

    print(f"Found {len(room_dirs)} room directories on disk.")

    # 3. Load excel database
    print("\nReading Excel database...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    db_products = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Rates":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        # Headers: Space | Code | Product | W | D | H | SH | MH | Dimension ...
        for row in rows[1:]:
            if len(row) < 3 or row[2] is None:
                continue
            name_str = str(row[2]).strip()
            code_str = str(row[1]).strip() if row[1] is not None else ""
            db_products.append({
                "sheet": sheet_name,
                "space": str(row[0]).strip() if row[0] is not None else "",
                "code": code_str,
                "name": name_str,
                "norm_name": normalize_name(name_str),
                "norm_code": normalize_name(code_str),
                "w": row[3],
                "d": row[4],
                "h": row[5],
                "sh": row[6],
                "mh": row[7],
                "dim_str": str(row[8]).strip() if len(row) > 8 and row[8] is not None else ""
            })
    print(f"Loaded {len(db_products)} product rows from Excel.")

    # 3b. Load rates database from the same unified database
    rates_db = {}
    norm_rates_db = {}
    
    # First, populate from the product sheets (using the Rate column in index 9, if it exists)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Rates":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        for row in rows[1:]:
            if len(row) < 3 or row[2] is None:
                continue
            name_str = str(row[2]).strip()
            code_str = str(row[1]).strip() if row[1] is not None else ""
            rate_val = row[9] if len(row) > 9 and row[9] is not None else None
            if rate_val is not None:
                rates_db[name_str] = rate_val
                rates_db[code_str] = rate_val
                norm_rates_db[normalize_name(name_str)] = name_str
                if code_str:
                    norm_rates_db[normalize_name(code_str)] = code_str

    # Then, overwrite/supplement with the dedicated "Rates" sheet if present.
    # (This stays as a fallback for any product not yet in the rate reference file.)
    if "Rates" in wb.sheetnames:
        print("Reading Rates sheet from unified database...")
        ws_rates = wb["Rates"]
        for row in ws_rates.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                key_str = str(row[0]).strip()
                rates_db[key_str] = row[1]
                norm_rates_db[normalize_name(key_str)] = key_str
        print(f"Loaded {len(rates_db)} rate entries from unified database.")

    # 3c. Authoritative rate source: the standalone pricing workbook's "Sheet5"
    # (canonical SKU name in col A, live Item Rate in col B — both formula-linked
    # to the TeakCosting rollup). Loaded LAST so it overrides the fallbacks above,
    # making the pricing file the single source of truth for rates.
    rates_ref_rel = project.get("xlsx", {}).get("rates_ref")
    if rates_ref_rel:
        rates_ref_path = os.path.normpath(os.path.join(proj_dir, rates_ref_rel))
        if os.path.exists(rates_ref_path):
            print(f"\nReading authoritative rates from: {os.path.basename(rates_ref_path)}")
            rwb = openpyxl.load_workbook(rates_ref_path, data_only=True)
            if "Sheet5" in rwb.sheetnames:
                loaded = 0
                for row in rwb["Sheet5"].iter_rows(min_row=2, values_only=True):
                    name, rate = (row[0] if len(row) > 0 else None), (row[1] if len(row) > 1 else None)
                    if name and isinstance(rate, (int, float)):
                        key_str = str(name).strip()
                        rates_db[key_str] = rate
                        norm_rates_db[normalize_name(key_str)] = key_str
                        loaded += 1
                print(f"Loaded {loaded} authoritative rates from pricing file (overrides DB).")
            else:
                print("  Warning: 'Sheet5' not found in pricing file; using DB rates only.")
        else:
            print(f"  Warning: rates_ref not found at {rates_ref_path}; using DB rates only.")

    # 5. Room sequence config comes from the project (projects.py)

    # Custom rate-item mapping comes from the project (projects.py)
    norm_custom_rates = {normalize_name(k): v for k, v in custom_rates_mapping.items()}

    def make_keyplan_svg(room_id):
        # Small key-plan thumbnail: the room's own floor plan with its area highlighted.
        fl = room_floor.get(room_id)
        if not fl:
            return ""   # room not placed on any floor plan → omit the thumbnail
        vb = fl.get("viewbox", "0 0 1980 1980")
        _, _, vw, vh = vb.split()
        img = f"{proj_base}/{fl['image']}"
        hl = fl["hotspots"].get(room_id, '')
        rect = (f'<rect {hl} fill="#a27b5c" fill-opacity="0.4" stroke="#a27b5c" stroke-width="15" />'
                if hl else '')
        return f"""<svg viewBox="{vb}" fill="none" xmlns="http://www.w3.org/2000/svg">
                      <image href="{img}" x="0" y="0" width="{vw}" height="{vh}" />
                      {rect}
                    </svg>"""

    def build_floor_slider(prefix, floors, slide_body_fn):
        """Generic one-floor-per-slide horizontal slider (fits one viewport, no
        scroll — paged via arrows/dots instead). Used for both the Plan Key and
        the Furniture Layout, wherever a project has more than one floor.
        slide_body_fn(floor) -> inner HTML for that floor's slide.
        Returns (nav_html, slider_html, dots_html, is_multi)."""
        multi = len(floors) > 1
        slides = ""
        dots = ""
        for i, fl in enumerate(floors):
            label = fl.get("label") or ""
            slides += f"""        <div class="floor-slide" data-label="{label}">
{slide_body_fn(fl)}
        </div>\n"""
            dots += f'<button class="floor-dot{" active" if i == 0 else ""}" onclick="floorGo(\'{prefix}\', {i})" aria-label="{label}"></button>'

        nav = f"""
        <div class="floor-slider-nav">
          <span class="floor-slider-label" id="{prefix}-slider-label">{floors[0].get("label", "") if floors else ""}</span>
          <button class="floor-arrow" onclick="floorSlide('{prefix}', -1)" aria-label="Previous floor">&#8249;</button>
          <button class="floor-arrow" onclick="floorSlide('{prefix}', 1)" aria-label="Next floor">&#8250;</button>
        </div>""" if multi else ""

        slider = f"""<div class="floor-slider">
        <div class="floor-track" id="{prefix}-track">
{slides}        </div>
      </div>"""

        dots_row = f'\n      <div class="floor-dots" data-dots="{prefix}">{dots}</div>' if multi else ""
        return nav, slider, dots_row, multi

    # 7. Generate Index Page HTML (Using the Image on the left side)
    def generate_index_page():
        room_count = len([r for r in rooms_config if r["id"] != "room-materials"])
        list_rows = ""
        num = 1
        for r_conf in rooms_config:
            rid = r_conf["id"]
            if rid == "room-materials":
                continue  # Exclude Materials & Moodboard from the index list
            num += 1
            idx_str = f"{num:02d}"
            title = r_conf["title"]
            label = r_conf["label"]
            
            list_rows += f"""          <!-- Row: {title} -->
          <div class="index-row" id="index-row-{rid}" onclick="scrollToRoom('{rid}')" onmouseover="highlightRoom('{rid}')" onmouseout="unhighlightRoom('{rid}')">
            <div class="index-row-left">
              <span class="index-row-num">{idx_str}</span>
              <span class="index-row-name">{title}</span>
            </div>
            <span class="index-row-label">{label}</span>
          </div>\n"""

        # One interactive plan SVG per floor (Umang: one; Modern Times: GF + FF),
        # paged with the shared floor-slider so it fits one viewport with no scroll.
        def _plan_slide(fl):
            vb = fl.get("viewbox", "0 0 1980 1980")
            _, _, vw, vh = vb.split()
            img = f"{proj_base}/{fl['image']}"
            rects = ""
            for r_conf in rooms_config:
                rid = r_conf["id"]
                attrs = fl["hotspots"].get(rid)
                if not attrs:
                    continue
                rects += f"""              <!-- {r_conf['title']} -->
              <rect class="svg-room-rect" id="svg-rect-{rid}" {attrs} onclick="scrollToRoom('{rid}')" onmouseover="highlightRoom('{rid}')" onmouseout="unhighlightRoom('{rid}')" />\n"""
            return f"""          <svg class="blueprint-svg-large" viewBox="{vb}" xmlns="http://www.w3.org/2000/svg" style="border-radius: var(--radius-card); border: var(--border-width) solid var(--line);">
            <image href="{img}" x="0" y="0" width="{vw}" height="{vh}" />
            <g class="svg-rooms-group">
{rects}            </g>
          </svg>"""

        plan_nav, plan_slider, plan_dots, plan_multi = build_floor_slider("plan", keyplan_floors, _plan_slide)
        blueprint_class = "blueprint-large-container blueprint-multi" if plan_multi else "blueprint-large-container"

        return f"""    <!-- ==========================================================================
         ROOM 01: INDEX PAGE (PLAN KEY)
         ========================================================================== -->
    <main class="room-section index-section" id="room-00" data-room-title="Plan Key">

      <!-- Left Column: Title + Blueprint Image -->
      <section class="left-column">
        <div class="layout-header-row">
          <div class="space-title-container">
            <span class="space-number">01</span>
            <h1 class="space-title">Plan Key</h1>
          </div>{plan_nav}
        </div>
        <div class="{blueprint_class}" style="padding: 0; overflow: hidden; background: none; border: none; box-shadow: none;">
{plan_slider}
        </div>{plan_dots}
      </section>

      <!-- Right Column: List Key Table -->
      <section class="right-column index-right-column" style="overflow-y: hidden;">
        <div class="index-meta-header">
          <span class="index-label">01 · PLAN KEY</span>
          <h2 class="index-heading">{room_count} rooms,<br>one walkthrough.</h2>
        </div>

        <div class="index-list-container">
{list_rows}        </div>
      </section>
    </main>\n"""

    # 8. Loop through all 11 rooms and build their HTML content
    print("\nProcessing room folders and compiling cards...")
    html_blocks = []
    js_room_renders = {}
    active_indices = {"room-00": 0}

    # Add Index page block first
    html_blocks.append(generate_index_page())

    # Default SVG Layout details for room views
    layout_svg = """<svg viewBox="0 0 100 100" fill="none" xmlns="http://www.w3.org/2000/svg">
                      <!-- Outer Walls -->
                      <path d="M10 10 H90 V90 H10 Z" stroke="#D8D2C8" stroke-width="1.5" />
                      <path d="M10 10 V90 M90 10 V90" stroke="#6B6258" stroke-width="1" />
                      <!-- Bathroom Divider -->
                      <path d="M68 10 V90" stroke="#6B6258" stroke-width="1" stroke-dasharray="2 2" />
                      <!-- Bedroom door swing -->
                      <path d="M10 75 A15 15 0 0 1 25 90" stroke="#a27b5c" stroke-width="0.75" />
                      <line x1="25" y1="90" x2="25" y2="75" stroke="#a27b5c" stroke-width="0.75" />
                      <!-- Bed outline (Four Poster) -->
                      <rect x="25" y="30" width="34" height="40" fill="none" stroke="#a27b5c" stroke-width="1" />
                      <!-- Pillows -->
                      <rect x="29" y="32" width="11" height="7" rx="1" fill="none" stroke="#6B6258" stroke-width="0.75" />
                      <rect x="44" y="32" width="11" height="7" rx="1" fill="none" stroke="#6B6258" stroke-width="0.75" />
                      <!-- Side tables -->
                      <rect x="17" y="30" width="8" height="8" fill="none" stroke="#6B6258" stroke-width="0.75" />
                      <rect x="59" y="30" width="8" height="8" fill="none" stroke="#6B6258" stroke-width="0.75" />
                      <!-- Study Table and Chair -->
                      <rect x="25" y="80" width="22" height="10" fill="none" stroke="#6B6258" stroke-width="0.75" />
                      <circle cx="36" cy="74" r="3.5" fill="none" stroke="#6B6258" stroke-width="0.75" />
                    </svg>"""

    for r_idx, r_conf in enumerate(rooms_config):
        rid = r_conf["id"]
        folder = r_conf["folder"]
        title = r_conf["title"]
        quote = r_conf["quote"]
        # Materials & Moodboard is shared across every project — built once by
        # build_materials_page(), not repeated inside each project's walkthrough.
        if rid == "room-materials":
            continue
        r_idx_num = r_idx + 1
        idx_str = f"{r_idx_num:02d}"
        active_indices[rid] = 0

        rpath = os.path.join(proj_dir, folder)

        # Determine layout image details from the project config (web path from root)
        layout_info = layout_images_cfg.get(rid)
        if layout_info:
            layout_img_path = f"{proj_base}/{layout_info['file']}"
            layout_img_title = layout_info["title"]
        else:
            layout_img_path = ""
            layout_img_title = ""
        
        # Compress any newly-added PNGs to JPEG before scanning (see optimize_web_image).
        optimize_images_in(rpath, ROOM_RENDER_MAX_W)
        prod_dir = os.path.join(rpath, "Products")
        optimize_images_in(prod_dir, PRODUCT_IMAGE_MAX_W)

        # Load main renders (alphabetical, unless the room folder has an order.txt)
        renders = list_images_ordered(rpath)

        # Load products (alphabetical, unless Products/ has an order.txt)
        products = []
        if os.path.exists(prod_dir):
            UPHOLSTERY_LABELS = {"FB": "Fabric", "LE": "Leather", "CN": "Cane"}
            for pitem in list_images_ordered(prod_dir):
                img_norm = normalize_name(pitem)
                best_match = None
                best_score = 0

                # Material tag from the internal [SIZE-UP-OPT] code, e.g. "Chair 44 C1
                # [1S-FB-N].jpg" -> UP="FB" -> "Teak + Fabric". XX (no upholstery) -> "Teak".
                material_tag = "Teak"
                up_match = re.search(r'\[[^\]-]*-([A-Za-z]{2})-[^\]]*\]', pitem)
                if up_match:
                    up_label = UPHOLSTERY_LABELS.get(up_match.group(1).upper())
                    if up_label:
                        material_tag = f"Teak + {up_label}"

                for db_p in db_products:
                    db_norm_name = db_p["norm_name"]
                    db_norm_code = db_p["norm_code"]

                    if img_norm == db_norm_name or img_norm == db_norm_code:
                        best_match = db_p
                        break
                    if img_norm in db_norm_name or db_norm_name in img_norm:
                        score = min(len(img_norm), len(db_norm_name)) / max(len(img_norm), len(db_norm_name))
                        if score > best_score:
                            best_score = score
                            best_match = db_p

                dim_label = ""
                if best_match:
                    w, d, h = best_match["w"], best_match["d"], best_match["h"]
                    sh, mh = best_match["sh"], best_match["mh"]
                    dim_parts = []
                    if w and d and h:
                        dim_parts.append(f"{w} x {d} x {h} cm (H)")
                    elif best_match["dim_str"]:
                        dim_parts.append(best_match["dim_str"])
                    if sh:
                        dim_parts.append(f"SH-{sh}cm")
                    if mh:
                        dim_parts.append(f"MH-{mh}cm")
                    dim_label = " &nbsp;·&nbsp; ".join(dim_parts)
                    prod_name = best_match["name"]
                else:
                    prod_name = re.sub(r'\.png$|\.jpg$|\.jpeg$', '', pitem)
                    prod_name = re.sub(r'\s*\[.*?\]', '', prod_name) # Drop internal-code bracket from client-facing name
                    prod_name = re.sub(r'\bv2\b|\bv1\b|\bfull image\b', '', prod_name).strip()
                    dim_label = "" # Leave empty if missing

                # Map rates
                rate_key = None
                for look_key in [img_norm, best_match["norm_name"] if best_match else None, best_match["norm_code"] if best_match else None]:
                    if look_key and look_key in norm_custom_rates:
                        rate_key = norm_custom_rates[look_key]
                        break

                if not rate_key:
                    for look_key in [img_norm, best_match["norm_name"] if best_match else None, best_match["norm_code"] if best_match else None]:
                        if look_key and look_key in norm_rates_db:
                            rate_key = norm_rates_db[look_key]
                            break

                rate_val = rates_db.get(rate_key) if rate_key else None

                products.append({
                    "img_path": f"{proj_base}/{folder}/Products/{pitem}",
                    "name": prod_name,
                    "dimensions": dim_label,
                    "rate": rate_val,
                    "material_tag": material_tag
                })

        # JS Render List
        js_renders = []
        for r in renders:
            js_renders.append({
                "src": f"{proj_base}/{folder}/{r}",
                "title": f"{title} Render",
                "sub": copy["render_sub"]
            })
        js_room_renders[rid] = js_renders

        # Slider HTML
        if len(renders) > 1:
            slider_html = f"""          <div class="main-render-wrapper" onclick="openLightbox('{rid}', activeRenderIndex['{rid}'])">
            <!-- Navigation Arrows on Slider -->
            <button class="slider-arrow slider-prev" onclick="navigateRender('{rid}', -1); event.stopPropagation();">&#8249;</button>
            <button class="slider-arrow slider-next" onclick="navigateRender('{rid}', 1); event.stopPropagation();">&#8250;</button>
            
            <img src="{proj_base}/{folder}/{renders[0]}" alt="{title} Render 1" class="main-render-image">
          </div>"""
        elif len(renders) == 1:
            slider_html = f"""          <div class="main-render-wrapper" onclick="openLightbox('{rid}', 0)">
            <img src="{proj_base}/{folder}/{renders[0]}" alt="{title} Render" class="main-render-image">
          </div>"""
        else:
            slider_html = f"""          <div class="main-render-wrapper">
            <div style="width:100%; height:100%; display:flex; flex-direction:column; justify-content:center; align-items:center; background:var(--cream); color:var(--muted);">
              <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.2" style="margin-bottom:10px;">
                <rect x="3" y="3" width="18" height="18" rx="2" />
                <circle cx="8.5" cy="8.5" r="1.5" />
                <path d="M21 15l-5-5L5 21" />
              </svg>
              <span style="font-family:var(--font-display); font-size: 0.95rem; font-weight:600; letter-spacing:0.05em; text-transform:uppercase;">{title} Renders Placeholder</span>
            </div>
          </div>"""

        # Products cards HTML
        products_html = ""
        if products:
            for p_idx, p in enumerate(products):
                p_name_escaped = html.escape(p["name"])
                
                rate_str = ""
                rate_lbl = ""
                if p.get("rate"):
                    material_tag = p.get("material_tag", "Teak")
                    rate_str = f"  —  Starting at ₹ {p['rate']:,} ({material_tag})"
                    rate_lbl = f"""\n                  <span style="display: flex; flex-direction: column; align-items: flex-end; gap: 2px;">
                    <span class="product-card__price" style="font-family: var(--font-display); font-size: 0.82rem; font-weight: 600; color: var(--chamoisee); white-space: nowrap;"><span style="font-family: var(--font-body); font-size: 0.6rem; font-weight: 500; color: var(--muted); opacity: 0.85;">Starting at </span>₹ {p['rate']:,}</span>
                    <span style="font-family: var(--font-body); font-size: 0.66rem; font-weight: 500; color: var(--muted); opacity: 0.75; white-space: nowrap;">{material_tag}</span>
                  </span>"""

                if p["dimensions"]:
                    details_label = p["dimensions"]
                    lightbox_details = f"{p['dimensions'].replace('&nbsp;·&nbsp;', ' — ')}{rate_str}"
                else:
                    details_label = ""
                    lightbox_details = f"{rate_str.lstrip(' — ')}"
                
                lightbox_details_escaped = html.escape(lightbox_details)
                
                products_html += f"""            <!-- Product Card: {p['name']} -->
            <div class="product-card">
              <div class="product-image-wrapper" onclick="openLightbox('{rid}-prod-{p_idx}', 0, '{p['img_path']}', '{p_name_escaped}', '{lightbox_details_escaped}')">
                <img src="{p['img_path']}" alt="{p_name_escaped}" class="product-image">
              </div>
              <div class="product-card__body">
                <div class="product-card__text">
                  <h3 class="product-card__name" style="margin-bottom: 0;">{p['name']}</h3>
                  <span class="product-card__material">{details_label}</span>
                </div>{rate_lbl}
              </div>
            </div>\n"""
        else:
            products_html = f"""            <div style="grid-column: span 2; font-family:var(--font-body); font-size:0.85rem; color:var(--muted); font-style:italic; padding: 40px 0; text-align:center; border: 1px dashed var(--line); border-radius: var(--radius-card); background: var(--light-secondary);">
              Custom bespoke furniture. Details TBD.
            </div>\n"""

        keyplan_svg = make_keyplan_svg(rid)

        # Maps column: detailed room layout and/or key-plan thumbnail — only when
        # those assets exist. If neither, the quote spans the row full-width.
        map_cards = ""
        if layout_img_path:
            map_cards += f"""          <div class="map-placeholder-card" title="Detailed Room Layout" onclick="openLightbox(null, 0, '{layout_img_path}', '{layout_img_title}', 'Room wise layout')">
            <div class="map-svg-wrapper">
              <img src="{layout_img_path}" alt="{layout_img_title}" class="map-layout-image">
            </div>
          </div>\n"""
        if keyplan_svg:
            map_cards += f"""          <div class="map-keyplan-large" title="Key House Plan — click to return to Plan Key" onclick="scrollToRoom('room-00')">
            <div class="map-svg-wrapper">
              {keyplan_svg}
            </div>
          </div>\n"""

        if map_cards:
            meta_header = f"""        <div class="room-meta-header">
          <blockquote class="quote-container">
            "{quote}"
          </blockquote>

          <div class="maps-container">
{map_cards}          </div>
        </div>"""
        else:
            meta_header = f"""        <div class="room-meta-header" style="grid-template-columns: 1fr;">
          <blockquote class="quote-container">
            "{quote}"
          </blockquote>
        </div>"""

        room_block = f"""    <!-- ==========================================================================
         ROOM {idx_str}: {title.upper()}
         ========================================================================== -->
    <main class="room-section" id="{rid}" data-room-title="{title}">

      <!-- Left Column: Title + Main Renders -->
      <section class="left-column">
        <div class="space-title-container">
          <span class="space-number">{idx_str}</span>
          <h1 class="space-title">{title}</h1>
        </div>

        <div class="showcase-container">
{slider_html}
        </div>
      </section>

      <!-- Right Column: Quote + Maps + Products Grid -->
      <section class="right-column">
        <!-- Top Row: Quote & Skeletal Maps -->
{meta_header}

        <!-- Bottom Row: Products Grid -->
        <div class="products-section-container">
          <h2 class="section-label">Furniture Used</h2>
          
          <div class="products-grid">
{products_html}          </div>
        </div>
      </section>
    </main>\n"""
        html_blocks.append(room_block)

    # 8b. Generate Customisation Page (last section)
    def generate_customisation_page():
        return f"""    <!-- ==========================================================================
         CUSTOMISATION: MADE FOR YOU
         ========================================================================== -->
    <main class="room-section custom-section-new" id="room-custom" data-room-title="Customisation">

      <!-- Top Row: Split Header -->
      <div class="custom-header-row">
        <div class="custom-header-left">
          <div class="custom-brand-group">
            <img src="Noku mark.png" alt="N" class="custom-brand-logo">
            <div class="custom-brand-text">
              <span class="custom-brand-title">CUSTOMISATION</span>
            </div>
          </div>
          <h1 class="custom-hero-heading">Every piece, tailored<br>to the residence.</h1>
        </div>
        
        <div class="custom-header-right">
          <div class="custom-chapter-tag">Chapter Four &nbsp;·&nbsp; Tailoring</div>
          <blockquote class="custom-description">
            {copy['customisation_intro']}
          </blockquote>
        </div>
      </div>

      <!-- Bottom Row: 4-Column Cards Grid -->
      <div class="custom-cards-row">
        <div class="custom-cards-grid-new">

          <!-- Card 1: Wood -->
          <div class="custom-card-new" title="Wood Customisations">
            <div class="custom-card-new__image-wrapper">
              <img src="Materials/Wood/Teak.png" alt="Wood Options" class="custom-card-new__image">
            </div>
            <div class="custom-card-new__body">
              <span class="custom-card-new__tag">Wood</span>
              <h3 class="custom-card-new__title">Teak, Reclaimed Teak, White Ash.</h3>
              <p class="custom-card-new__desc">Mix-and-match across pieces. Reclaimed available subject to availability. Prices quoted throughout are starting prices in teak — other wood options priced on request.</p>
            </div>
          </div>

          <!-- Card 2: Upholstery -->
          <div class="custom-card-new" title="Upholstery Options">
            <div class="custom-card-new__image-wrapper">
              <img src="Materials/Fabric/Rubik Linen.jpg" alt="Upholstery Options" class="custom-card-new__image">
            </div>
            <div class="custom-card-new__body">
              <span class="custom-card-new__tag">Upholstery</span>
              <h3 class="custom-card-new__title">Fabric or Italian Leather.</h3>
              <p class="custom-card-new__desc">Library of 40+ swatches across our partners &mdash; Vagabond, D&rsquo;Decor, Rubik.</p>
            </div>
          </div>

          <!-- Card 3: Dimensions -->
          <div class="custom-card-new custom-card-new--dim" title="Dimension Customisations">
            <div class="custom-card-new__image-wrapper custom-card-new__image-wrapper--dim">
              <svg viewBox="0 0 120 120" fill="none" xmlns="http://www.w3.org/2000/svg" class="custom-dim-svg">
                <rect x="35" y="25" width="50" height="50" stroke="#a27b5c" stroke-width="1.2"/>
                <line x1="35" y1="25" x2="85" y2="75" stroke="#a27b5c" stroke-width="1.0"/>
                <line x1="85" y1="25" x2="35" y2="75" stroke="#a27b5c" stroke-width="1.0"/>
                <text x="60" y="96" text-anchor="middle" font-family="Josefin Sans, sans-serif" font-size="8.5" fill="#a27b5c" letter-spacing="1">DIM &plusmn;10%</text>
              </svg>
            </div>
            <div class="custom-card-new__body">
              <span class="custom-card-new__tag">Dimensions</span>
              <h3 class="custom-card-new__title">Scale to your space.</h3>
              <p class="custom-card-new__desc">Up to &plusmn;10% on most pieces with no tooling surcharge. Larger changes treated as a new study.</p>
            </div>
          </div>

        </div><!-- /custom-cards-grid-new -->
      </div><!-- /custom-cards-row -->
    </main>\n"""

    html_blocks.append(generate_customisation_page())

    # 8c. Generate Thank You Page (last section)
    def generate_thank_you_page():
        return f"""    <!-- ==========================================================================
         SECTION 14: THANK YOU (dark)
         ========================================================================== -->
    <section class="room-section landing-thankyou-section" id="thank-you" data-room-title="Thank You">
      <!-- Background Collage Panel / Vignette -->
      <div class="ty-bg-panel">
        <img src="Moodboard - of stillness/01.jpg" alt="Thank You background">
      </div>
      <div class="ty-bg-vignette"></div>

      <!-- Tickers -->
      <div class="ty-side-ticker"><span>{copy['thankyou_side_ticker']}</span></div>
      <div class="ty-edition-mark"><span>Noku Studio &nbsp;·&nbsp; Edition 2026</span></div>

      <!-- Header Nav -->
      <nav class="ty-nav" aria-label="Thank you page navigation">
        <div class="ty-nav__brand">
          <img src="Noku mark.png" alt="Noku mark" class="ty-nav__logo">
          <span class="ty-nav__wordmark">Noku Studio</span>
        </div>
        <span class="ty-nav__label">END OF PRESENTATION</span>
      </nav>

      <!-- Stage Content -->
      <div class="ty-stage">
        <div class="ty-eyebrow">
          <span class="ty-eyebrow-line"></span>
          <span class="ty-eyebrow-text">Good Earth Kochi</span>
          <span class="ty-eyebrow-line"></span>
        </div>
        <h1 class="ty-title">Thank You.</h1>
        <div class="ty-divider"></div>
        <p class="ty-descriptor">
          {copy['thankyou_descriptor']}
        </p>
        
        <div class="ty-ctas">
{cross_ctas}        </div>
      </div>

      <!-- Footer -->
      <footer class="ty-footer">
        <div class="ty-footer__left">
          <span class="ty-footer__credit">Designed by Noku Studio</span>
          <span class="ty-footer__client">{copy['thankyou_footer_client']}</span>
        </div>
        <span class="ty-pagination">30 &nbsp;/&nbsp; 30</span>
      </footer>
    </section>"""

    # 9. Furniture Layout section (unnumbered) — a horizontal slider across floors.
    print(f"\nMerging blocks into {project['output']}...")

    def _furn_slide(fl):
        f_img = f"{proj_base}/{fl['file']}"
        f_title = fl["title"]
        return f"""          <div class="layout-image-container-centered" onclick="openLightbox(null, 0, '{f_img}', '{f_title}')">
            <img src="{f_img}" alt="{f_title}" class="furniture-layout-image">
          </div>"""

    furn_nav, furn_slider, furn_dots_row, furn_multi = build_floor_slider("furn", furniture_floors, _furn_slide)

    if furn_multi:
        left_arrow = f"""<button class="floor-arrow floor-arrow--prev" onclick="floorSlide('furn', -1)" aria-label="Previous floor" style="position: absolute; left: 15px; top: 50%; transform: translateY(-50%); z-index: 10;">&#8249;</button>"""
        right_arrow = f"""<button class="floor-arrow floor-arrow--next" onclick="floorSlide('furn', 1)" aria-label="Next floor" style="position: absolute; right: 15px; top: 50%; transform: translateY(-50%); z-index: 10;">&#8250;</button>"""
        
        furn_nav_clean = f"""
        <div class="floor-slider-nav">
          <span class="floor-slider-label" id="furn-slider-label">{furniture_floors[0].get("label", "") if furniture_floors else ""}</span>
        </div>"""
        
        slider_wrapper = f"""<div class="furniture-slider-wrapper" style="position: relative; width: 100%; display: flex; align-items: center;">
        {left_arrow}
        {furn_slider}
        {right_arrow}
      </div>"""
    else:
        furn_nav_clean = furn_nav
        slider_wrapper = furn_slider

    furniture_layout_block = f"""    <!-- ==========================================================================
         FURNITURE LAYOUT: HORIZONTAL SLIDER ACROSS FLOORS (UNNUMBERED)
         ========================================================================== -->
    <main class="room-section furniture-layout-section{' furniture-multi' if furn_multi else ''}" id="room-layout" data-room-title="Furniture Layout">

      <!-- Top Title Bar -->
      <div class="layout-header-row">
        <div class="space-title-container">
          <h1 class="space-title">Furniture Layout</h1>
        </div>{furn_nav_clean}
      </div>

      {slider_wrapper}{furn_dots_row}
    </main>"""

    # Cross-project hand-off on the Thank You page: continue to the other project(s), or the hub
    others = [p for p in PROJECTS if p["id"] != project["id"]]
    cross_ctas = "".join(f"""          <a class="ty-btn ty-btn--primary" href="{o['output']}">
            <span>Continue to {o['short_name']}</span>
            <span class="ty-btn__arrow">→</span>
          </a>\n""" for o in others)
    cross_ctas += """          <a class="ty-btn" href="index.html">
            <span>All Projects</span>
            <span class="ty-btn__arrow">↗</span>
          </a>\n"""

    if project.get("status") == "preview":
        # Furnishing study underway — a single elegant preview section replaces the
        # room walkthrough until real content lands (drop the flag in projects.py then).
        preview_block = f"""    <!-- ==========================================================================
         PREVIEW: FURNISHING STUDY UNDERWAY
         ========================================================================== -->
    <main class="room-section" id="room-preview" data-room-title="In Design"
      style="grid-template-columns: 1fr; align-content: center; justify-items: center; text-align: center;">
      <div style="max-width: 620px; display: flex; flex-direction: column; align-items: center; gap: 26px; padding: 60px 20px;">
        <img src="Noku mark.png" alt="Noku Studio" style="height: 34px; width: auto; opacity: 0.9;">
        <span class="index-label">Furnishing Study Underway</span>
        <h1 class="space-title" style="font-size: clamp(2rem, 4.5vw, 3.4rem);">The walkthrough is<br>being composed.</h1>
        <p style="font-family: var(--font-body); font-size: 1rem; line-height: 1.8; color: var(--muted);">
          Noku Studio is composing the room-by-room furnishing study for
          <strong style="color: var(--ink); font-weight: 400;">{project['name']}</strong> —
          renders, product selections, dimensions and rates will appear here.
          The shared materials palette and moodboard already apply to this residence.
        </p>
        <div style="display: flex; gap: 28px; flex-wrap: wrap; justify-content: center;">
          <a href="index.html#materials-moodboard" style="font-family: var(--font-display); font-size: 0.78rem; letter-spacing: 0.12em; text-transform: uppercase; color: var(--chamoisee); text-decoration: none; border-bottom: 1px solid var(--chamoisee); padding-bottom: 3px;">Shared Materials &amp; Moodboard →</a>
          <a href="{others[0]['output']}" style="font-family: var(--font-display); font-size: 0.78rem; letter-spacing: 0.12em; text-transform: uppercase; color: var(--chamoisee); text-decoration: none; border-bottom: 1px solid var(--chamoisee); padding-bottom: 3px;">See the {others[0]['short_name']} walkthrough →</a>
        </div>
      </div>
    </main>\n"""
        ordered_blocks = [preview_block, generate_thank_you_page()]
    else:
        # Plan Key and Furniture Layout are included only when their images exist
        # (html_blocks[0] = index/plan key, html_blocks[1:] = rooms + customisation).
        lead = []
        if has_plan:
            lead.append(html_blocks[0])
        if has_furniture:
            lead.append(furniture_layout_block)
            active_indices["room-layout"] = 0
        ordered_blocks = lead + html_blocks[1:] + [generate_thank_you_page()]

    orig_html = merge_rooms_into_shell(shell_html, ordered_blocks, js_room_renders, active_indices)
    if orig_html is None:
        return

    # 10. Per-project tokens: title, meta description, embedded cover, hero background
    name = project["name"]
    orig_html = set_shell_tokens(
        orig_html,
        title=f"{name} — Room Showcase",
        meta_description=copy["meta_description"],
        brand_name=name,
    )
    orig_html = set_embedded_cover(orig_html, project["hero"], side_ticker_project=name)

    # 11. Write the generated walkthrough for this project
    out_path = os.path.join(base_dir, project["output"])
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(orig_html)

    print(f"Success! {project['output']} written.")


def main():
    print("Starting Noku Pitch Webpage Auto-Updater...")
    base_dir = os.path.dirname(os.path.abspath(__file__))

    template_path = os.path.join(base_dir, TEMPLATE_FILE)
    if not os.path.exists(template_path):
        print(f"Error: template not found at: {template_path}")
        return
    with open(template_path, 'r', encoding='utf-8') as f:
        shell_html = f.read()

    inject_materials_section(base_dir)

    for project in PROJECTS:
        build_project(project, shell_html, base_dir)

    print("\nAll projects built. Open index.html to pick a project.")


if __name__ == "__main__":
    main()
